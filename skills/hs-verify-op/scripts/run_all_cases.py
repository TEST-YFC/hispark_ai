#!/usr/bin/env python3
# coding: utf-8
"""hs-verify-op FIXED harness — operator-agnostic. Run IN PLACE from the skill's scripts/.

This file is the single verification entry point described in SKILL.md. It is
operator-AGNOSTIC and must NOT be edited per operator: the ONLY thing you write per
operator is an `op_spec.py` (copy from operator_spec_template.py) supplying the test
cases and the model builders. The harness then runs, for each framework independently,
the strict internal step1->step5 pipeline and writes one Excel per framework.

  step1  build model + deterministic inputs + reference output (onnxruntime / tf.lite)
  step2  converter_lite -> micro C project        (bundled *_<target>.sh)
  step3  cmake + make benchmark                    (bundled *_<target>.sh)
  step4  write input .bin                          (here)
  step5  run benchmark (prints output tensors), then compute cosine in Python — UNIFORM
         for x86 AND riscv (same cosine_similarity, same thresholds). The benchmark only
         PRINTS tensors; it NEVER decides PASS/FAIL. There is exactly one cosine code path.

Usage
-----
    cd $MSLITE_OP_OUTPUT/<op>                    # default sits beside the mindspore-lite repo (src/mslite-op-output/ inside HiSpark.AI)
    source scripts/env_setup.sh                  # or: export MSLITE_PKG=...
    python <skill>/scripts/run_all_cases.py --spec scripts/op_spec.py \
        [--framework {onnx,tflite,all}] [--target {x86,riscv,all}]

Hard rules (enforced here, see SKILL.md Red Flags)
--------------------------------------------------
* Cosine values are ONLY ever the parsed output of a real benchmark run. The harness
  never invents a number; a path that cannot run is recorded as FAIL with its error.
* Reference outputs are computed by the harness from the model; the spec never supplies
  "expected" numbers. Thresholds are fixed (fp32>=0.999, INT8>=0.99) and not overridable
  to turn red green — only loosenable via explicit flags for debugging, never for sign-off.
* Missing toolchain -> the harness stops and reports; it does not fabricate PASS.
"""

import argparse
import datetime
import importlib.util
import json
import os
import re
import shlex
import shutil
import signal
import subprocess
import sys
from pathlib import Path

import numpy as np

SCRIPT_DIR = Path(__file__).parent.resolve()

DEFAULT_THRESHOLD_FP32 = 0.999
DEFAULT_THRESHOLD_INT8 = 0.99

FRAMEWORKS = {"onnx": ["onnx"], "tflite": ["tflite"], "all": ["onnx", "tflite"]}
TARGET_PATHS = {
    "x86":   ["x86_fp32"],
    "riscv": ["riscv_fp32", "riscv_int8"],
    "all":   ["x86_fp32", "riscv_fp32", "riscv_int8"],
}
# col label + PASS threshold per target path (shared by both framework tables)
PATH_META = {
    "x86_fp32":   {"col": "x86 fp32 余弦",   "thr": DEFAULT_THRESHOLD_FP32},
    "riscv_fp32": {"col": "riscv fp32 余弦", "thr": DEFAULT_THRESHOLD_FP32},
    "riscv_int8": {"col": "riscv INT8 余弦", "thr": DEFAULT_THRESHOLD_INT8},
}
# (framework, path) -> bundled driver script + cfg template
DRIVER = {
    ("onnx",   "x86_fp32"):   ("onnx_x86.sh",    "micro_x86.cfg"),
    ("onnx",   "riscv_fp32"): ("onnx_riscv.sh",  "micro_riscv.cfg"),
    ("onnx",   "riscv_int8"): ("onnx_riscv.sh",  "micro_riscv_quant.cfg"),
    ("tflite", "x86_fp32"):   ("tflite_x86.sh",  "micro_x86.cfg"),
    ("tflite", "riscv_fp32"): ("tflite_riscv.sh", "micro_riscv.cfg"),
    ("tflite", "riscv_int8"): ("tflite_riscv.sh", "micro_riscv_quant.cfg"),
}
MODEL_EXT = {"onnx": "onnx", "tflite": "tflite"}


# --------------------------------------------------------------------------- #
# Environment / dependencies
# --------------------------------------------------------------------------- #

def ensure(mod, pkg=None):
    """Lazily pip-install a dependency into the current interpreter (Tsinghua mirror)."""
    try:
        __import__(mod)
    except ImportError:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", pkg or mod,
             "-i", "https://pypi.tuna.tsinghua.edu.cn/simple"],
            check=True,
        )


def _check_pkg_freshness(pkg: str):
    """Refuse a stale extracted package. A rebuilt tarball next to an OLD extraction is the
    classic way to verify yesterday's converter_lite and reach a wrong verdict: rebuilding
    refreshes the tar.gz, but nothing forces a re-extract, and the stale binary keeps
    "working". Detection: tar restores file *mtimes* from the archive but cannot set
    *ctime* — so ctime(converter_lite) is the real extraction moment. A sibling tarball
    newer than that means the extraction predates the latest build -> refuse.
    Override with OP_VERIFY_ALLOW_STALE=1 only for deliberate archaeology."""
    if os.environ.get("OP_VERIFY_ALLOW_STALE") == "1":
        return
    pkg_path = Path(pkg).resolve()
    conv = pkg_path / "tools/converter/converter/converter_lite"
    candidates = [pkg_path.parent / (pkg_path.name + ".tar.gz"),
                  pkg_path.parent / "tmp" / (pkg_path.name + ".tar.gz")]
    for tb in candidates:
        if tb.is_file() and tb.stat().st_mtime > conv.stat().st_ctime:
            sys.exit(
                "[ERROR] MSLITE_PKG 解压包比其旁的构建产物 tar.gz 旧（解压发生在最近一次构建之前）。\n"
                f"        包:     {pkg_path}\n"
                f"        tar.gz: {tb}  (更新)\n"
                "        在旧 converter_lite 上跑出的结论不反映当前代码——这是历史上真实发生过的假结论来源。\n"
                "        重新解压后再跑（build_mslite.sh 构建成功后会自动解压；或手动 rm -rf 包目录后 tar xzf）。\n"
                "        确要用旧包对比历史行为时，OP_VERIFY_ALLOW_STALE=1 显式放行。")


def resolve_mslite_pkg(start: Path) -> str:
    """MSLITE_PKG from env, else auto-locate by walking up for the built toolchain.
    Either way the package must not be staler than its sibling tarball."""
    env = os.environ.get("MSLITE_PKG")
    if env and Path(env, "tools/converter/converter/converter_lite").is_file():
        _check_pkg_freshness(env)
        return env
    rel = "src/mindspore-lite/output/mindspore-lite-2.8.0-linux-x64"
    for parent in [start, *start.parents]:
        cand = parent / rel
        if (cand / "tools/converter/converter/converter_lite").is_file():
            _check_pkg_freshness(str(cand))
            return str(cand.resolve())
    sys.exit(
        "[ERROR] MSLITE_PKG 未设置且无法自动定位已构建的 MindSpore Lite。\n"
        "        先 `source scripts/env_setup.sh` 或 `export MSLITE_PKG=<toolchain>`，\n"
        "        且该路径下需存在 tools/converter/converter/converter_lite。不构建则停止，不伪造结果。"
    )


# --------------------------------------------------------------------------- #
# Spec loading
# --------------------------------------------------------------------------- #

REQUIRED_SPEC_ATTRS = ["OP_NAME", "ONNX_TEST_CASES", "TFLITE_TEST_CASES",
                       "build_onnx_model", "build_tflite_model", "make_inputs"]


def load_spec(spec_path: Path):
    if not spec_path.is_file():
        sys.exit(f"[ERROR] 算子 spec 文件不存在: {spec_path}\n"
                 f"        从 {SCRIPT_DIR / 'operator_spec_template.py'} 拷贝并填好后再运行。")
    mod_spec = importlib.util.spec_from_file_location("op_spec", spec_path)
    mod = importlib.util.module_from_spec(mod_spec)
    mod_spec.loader.exec_module(mod)
    missing = [a for a in REQUIRED_SPEC_ATTRS if not hasattr(mod, a)]
    if missing:
        sys.exit(f"[ERROR] op_spec.py 缺少必需定义: {missing}")
    if not hasattr(mod, "PARAM_COLUMNS"):
        mod.PARAM_COLUMNS = []
    # Optional: TFLITE_TARGET_BUILTIN (int) — the builtin number from existence
    # verification. When present, every built .tflite is unpacked and asserted to
    # actually contain it (see _assert_target_builtin).
    return mod


def load_spec_module(spec_path: Path):
    """Load an op_spec module without enforcing the full harness-run contract.

    Manual per-path judgement only needs OP_NAME / optional INT8_KERNEL_SYMBOL for the
    INT8 genuineness gate. Full runs still use load_spec(), which validates every required
    builder/case attribute before burning time.
    """
    mod_spec = importlib.util.spec_from_file_location("op_spec", spec_path)
    mod = importlib.util.module_from_spec(mod_spec)
    mod_spec.loader.exec_module(mod)
    return mod


# --------------------------------------------------------------------------- #
# Reference inference (ground truth — computed here, never supplied by the spec)
# --------------------------------------------------------------------------- #

def run_reference(framework, model_path, inputs):
    """Return (input_names, output_names, ref_outputs) for the built model.

    inputs: ordered list[np.ndarray] in model-input order (from spec.make_inputs).
    """
    if framework == "onnx":
        import onnxruntime as ort
        sess = ort.InferenceSession(str(model_path), providers=["CPUExecutionProvider"])
        in_names = [i.name for i in sess.get_inputs()]
        out_names = [o.name for o in sess.get_outputs()]
        feed = {name: arr for name, arr in zip(in_names, inputs)}
        outs = sess.run(None, feed)
        return in_names, out_names, [np.asarray(o) for o in outs]
    else:
        import tensorflow as tf
        interp = tf.lite.Interpreter(model_path=str(model_path))
        interp.allocate_tensors()
        in_det = interp.get_input_details()
        out_det = interp.get_output_details()
        for det, arr in zip(in_det, inputs):
            interp.set_tensor(det["index"], arr.astype(det["dtype"]))
        interp.invoke()
        in_names = [d["name"] for d in in_det]
        out_names = [d["name"] for d in out_det]
        outs = [interp.get_tensor(d["index"]) for d in out_det]
        return in_names, out_names, [np.asarray(o) for o in outs]


# --------------------------------------------------------------------------- #
# Benchmark output parsing / cosine
# --------------------------------------------------------------------------- #

def parse_benchmark_outputs(stdout):
    """The generated benchmark (x86 AND riscv — same benchmark.c) ALWAYS prints each output
    tensor as a 'name: ... Data:' header line followed by a comma-separated values line
    (PrintTensorHandle). Parse them so the harness computes cosine itself, identically for
    every target. This is the ONLY source of device output the harness trusts."""
    outs, lines = [], stdout.splitlines()
    for idx, line in enumerate(lines):
        if re.search(r"name:.*Data:\s*$", line.strip()):
            data_line = lines[idx + 1].strip() if idx + 1 < len(lines) else ""
            vals = [float(x) for x in data_line.split(",") if x.strip()]
            if vals:
                outs.append(np.array(vals, dtype=np.float32))
    return outs


def _tensor_output_name(idx, count):
    return f"output_{idx}.npy" if count > 1 else "output.npy"


def _load_numpy_outputs(out_dir: Path):
    return [np.load(p) for p in sorted(out_dir.glob("output*.npy"))]


def _clear_numpy_outputs(out_dir: Path):
    for p in out_dir.glob("output*.npy"):
        p.unlink()


def _write_judge_report(log_dir: Path, path_key: str, status: str, cos, msg: str):
    thr = PATH_META[path_key]["thr"]
    cos_s = "ERR" if cos is None else f"{cos:.6f}"
    report = (
        f"JUDGE: {path_key} {status} cos={cos_s} threshold>={thr}"
        + (f"  {msg}" if msg else "")
    )
    (log_dir / "judge.txt").write_text(report + "\n")
    return report


def judge_path_from_stdout(case_dir, path_key, stdout, stderr="", rc=0, spec=None, build_dir=None):
    """Refresh output*.npy from the latest benchmark stdout and compare against gt/.

    This is the single per-path judgement path used by both full harness runs and manual
    tc reruns through output/<path>/_run.sh. gt/output*.npy is stable; output/<path>/output*.npy
    is deliberately overwritten from the latest stdout so stale manual results cannot pass.
    """
    case_dir = Path(case_dir)
    log_dir = case_dir / "output" / path_key
    gt_dir = case_dir / "gt"
    log_dir.mkdir(parents=True, exist_ok=True)
    _clear_numpy_outputs(log_dir)

    bench = parse_benchmark_outputs(stdout)
    if not bench:
        msg = _err_msg(stdout, stderr, log_dir, rc)
        _write_judge_report(log_dir, path_key, "FAIL", None, msg)
        return "FAIL", None, msg

    for i, b in enumerate(bench):
        np.save(log_dir / _tensor_output_name(i, len(bench)), b)

    refs = _load_numpy_outputs(gt_dir)
    if not refs:
        msg = f"no gt output*.npy under {gt_dir}"
        _write_judge_report(log_dir, path_key, "FAIL", None, msg)
        return "FAIL", None, msg
    if len(bench) != len(refs):
        msg = f"output tensor count mismatch gt={len(refs)} vs device={len(bench)}"
        _write_judge_report(log_dir, path_key, "FAIL", None, msg)
        return "FAIL", None, msg

    cos = min(cosine_similarity(b, r) for b, r in zip(bench, refs))
    thr = PATH_META[path_key]["thr"]
    status = "PASS" if cos >= thr else "FAIL"
    msg = "" if status == "PASS" else f"cos {cos:.6f} < {thr}"

    if path_key == "riscv_int8" and spec is not None and build_dir is not None:
        int8_status, why = assert_int8_genuine(spec, Path(build_dir))
        if int8_status == "missing":
            msg = f"INT8_NOT_GENUINE: {why}" + (f"  [{msg}]" if msg else "")
            _write_judge_report(log_dir, path_key, "FAIL", cos, msg)
            return "FAIL", cos, msg
        note = "int8_genuine=yes" if int8_status == "genuine" else "int8_exempt=yes"
        msg = f"{msg}; {note}" if msg else note

    _write_judge_report(log_dir, path_key, status, cos, msg)
    return status, cos, msg


def cosine_similarity(a, b):
    """The ONE cosine in the harness — used for x86 and riscv alike. Mathematically defined
    for every input, so it NEVER returns NaN:
      * both vectors all-zero      -> 1.0  (both produced nothing -> they match)
      * exactly one vector all-zero -> 0.0  (a real mismatch -> FAIL)
    Do NOT add any `nan`/`inf` -> passing-value shortcut anywhere (parser, driver, or here).
    A NaN reaching a verdict means the cosine wasn't really computed; that is a bug to fix,
    never a PASS to fabricate. (See SKILL.md Red Flags.)"""
    a = np.asarray(a).flatten().astype(np.float64)
    b = np.asarray(b).flatten().astype(np.float64)
    na, nb = np.linalg.norm(a), np.linalg.norm(b)
    if na == 0 and nb == 0:
        return 1.0
    if na == 0 or nb == 0:
        return 0.0
    return float(np.dot(a, b) / (na * nb))


# --------------------------------------------------------------------------- #
# Shell driver execution
# --------------------------------------------------------------------------- #

def _quote_path(path):
    return shlex.quote(str(path))


def _write_rerunnable_driver(log_dir, build_dir, raw_driver, path_key, case_dir, spec_path=None):
    """Write the manual tc rerun wrapper and the raw converter/build/benchmark driver."""
    log_dir.mkdir(parents=True, exist_ok=True)
    driver_sh = log_dir / "_driver.sh"
    driver_sh.write_text(raw_driver)

    harness = Path(__file__).resolve()
    spec_arg = f" --spec {_quote_path(spec_path)}" if spec_path else ""
    wrapper = f"""#!/bin/bash
# Rerunnable hs-verify-op path wrapper. Safe to run from any CWD.
set -u
OUT_DIR="$(cd "$(dirname "$(realpath "${{BASH_SOURCE[0]}}")")" && pwd)"
CASE_DIR={_quote_path(Path(case_dir))}
BUILD_DIR={_quote_path(Path(build_dir))}
DRIVER_SH="$OUT_DIR/_driver.sh"
STDOUT_LOG="$OUT_DIR/stdout.log"
STDERR_LOG="$OUT_DIR/stderr.log"

mkdir -p "$BUILD_DIR" "$OUT_DIR"
(
  cd "$BUILD_DIR" || exit 1
  bash "$DRIVER_SH"
) > "$STDOUT_LOG" 2> "$STDERR_LOG"
driver_rc=$?

cat "$STDOUT_LOG"
cat "$STDERR_LOG" >&2

if [ "${{OP_VERIFY_SKIP_JUDGE:-0}}" = "1" ]; then
  exit "$driver_rc"
fi

python3 {_quote_path(harness)} --judge-case "$CASE_DIR" --judge-path {shlex.quote(path_key)} \\
  --stdout-log "$STDOUT_LOG" --stderr-log "$STDERR_LOG" --driver-rc "$driver_rc"{spec_arg}
exit $?
"""
    run_sh = log_dir / "_run.sh"
    run_sh.write_text(wrapper)
    return run_sh


def run_driver(framework, path_key, build_dir, log_dir, mslite_pkg, model_file, cfg_file, input_files,
               case_dir=None, spec_path=None):
    """Fill the bundled driver template, run it with cwd=build_dir, write logs to log_dir.

    Directory classification (model/convert/input/gt/output layout):
      build_dir = <case>/<path>/convert  — driver CWD; the *_micro build tree lands here
      log_dir   = <case>/<path>/output   — _run.sh / stdout.log / stderr.log land here
    The driver scripts create `onnx_x86_micro` (etc.) RELATIVE to their CWD, so pointing
    cwd at build_dir puts the conversion tree under convert/ with no script change."""
    driver_name, _ = DRIVER[(framework, path_key)]
    content = (SCRIPT_DIR / driver_name).read_text()
    repl = {
        "{MSLITE_PKG}": mslite_pkg,
        "{MODEL_FILE}": str(model_file),
        "{CFG_FILE}": str(cfg_file),
        "{INPUT_FILE}": ",".join(str(p) for p in input_files),
    }
    for k, v in repl.items():
        content = content.replace(k, v)
    build_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    if case_dir is None:
        case_dir = log_dir.parent.parent
    run_sh = _write_rerunnable_driver(log_dir, build_dir, content, path_key, case_dir, spec_path)
    # Backstop against a converter that hangs or spins on heap corruption instead of
    # aborting cleanly: bound every path and kill the WHOLE process group on timeout.
    # start_new_session makes bash the group leader, so converter_lite — its grandchild —
    # is killed too rather than orphaned. Override via OP_VERIFY_PATH_TIMEOUT (seconds);
    # this is a hang backstop, not a perf limit, so it defaults generously.
    timeout = int(os.environ.get("OP_VERIFY_PATH_TIMEOUT", "1200"))
    proc = subprocess.Popen(["bash", str(run_sh)], cwd=str(build_dir),
                            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
                            start_new_session=True)
    try:
        stdout, stderr = proc.communicate(timeout=timeout)
        rc = proc.returncode
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except ProcessLookupError:
            pass
        stdout, stderr = proc.communicate()
        stored_stdout = log_dir / "stdout.log"
        stored_stderr = log_dir / "stderr.log"
        stdout = (stdout or "") or (stored_stdout.read_text() if stored_stdout.is_file() else "")
        stderr = (stderr or "") or (stored_stderr.read_text() if stored_stderr.is_file() else "")
        stderr = stderr + (
            f"\n[ERR] path timed out after {timeout}s and was killed — converter likely hung "
            f"or spinning on heap corruption (raise OP_VERIFY_PATH_TIMEOUT if the path is "
            f"merely slow).\n")
        stored_stderr.write_text(stderr)
        rc = -9
    return stdout, stderr, rc


def make_cfg(path_key, cfg_tmpl_name, dst_dir, in_names, calib_dirs):
    """Materialize the concrete cfg for this path.

    Only riscv_int8 needs filling. Every model input gets its OWN calibration directory,
    so the cfg lists one `tensorName:calibDir` entry per input (TFLite Select/Where etc.
    have 3 inputs). This generalizes to any input count — no fixed first/last special-case.
        calibrate_path = name0:dir0,name1:dir1,...     (one entry per input)
        calibrate_size = number of calibration SAMPLES per input (we drop one .bin -> 1)
    """
    if path_key != "riscv_int8":
        return SCRIPT_DIR / cfg_tmpl_name  # x86 / riscv fp32 cfgs have no placeholders
    tmpl = (SCRIPT_DIR / cfg_tmpl_name).read_text()
    calibrate_path = ",".join(f"{name}:{d}" for name, d in zip(in_names, calib_dirs))
    tmpl = tmpl.replace("{CALIBRATE_PATH}", calibrate_path)
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / "micro_riscv_quant.cfg"
    dst.write_text(tmpl)
    return dst


def _tflite_builtin_codes(model_path):
    """Operator codes actually present in a built .tflite. The TF converter may NORMALIZE
    the op the builder asked for into a different builtin depending on shapes (real case:
    same-shape calls to a broadcast-select op get lowered to its non-broadcast sibling).
    A case whose model lacks the target builtin silently tests a DIFFERENT operator —
    its PASS/FAIL says nothing about the operator under verification."""
    from tensorflow.lite.python import schema_py_generated as s
    m = s.Model.GetRootAsModel(Path(model_path).read_bytes(), 0)
    return sorted({max(m.OperatorCodes(i).BuiltinCode(), m.OperatorCodes(i).DeprecatedBuiltinCode())
                   for i in range(m.OperatorCodesLength())})


def _assert_target_builtin(spec, framework, model_file):
    """Per-case enforcement of SKILL.md's 「目标 builtin 实证」 when the spec declares
    TFLITE_TARGET_BUILTIN (the builtin number hit during existence verification)."""
    target = getattr(spec, "TFLITE_TARGET_BUILTIN", None)
    if framework != "tflite" or target is None:
        return
    codes = _tflite_builtin_codes(model_file)
    if int(target) not in codes:
        raise RuntimeError(
            f"OP_MISMATCH: model builtins={codes}, target builtin {target} absent — "
            f"converter normalized the op into a different builtin; this case would test "
            f"the WRONG operator. Fix the case's shapes in op_spec.py so the target builtin "
            f"is emitted (do NOT accept the substitute builtin, do NOT delete the case).")


def _names_from_value_infos(value_infos):
    return [getattr(v, "name", str(v)) for v in value_infos]


def _initializer_inputs_for_case(spec, framework, tc):
    """Names of model inputs intentionally backed by initializers/constants instead of
    make_inputs() arrays. A spec may declare either INITIALIZER_INPUTS = {"onnx": [...]}
    or a callable INITIALIZER_INPUTS(framework, tc) for case-dependent optional inputs.
    This keeps native-integer ops honest: graph inputs are either fed dynamically or
    explicitly documented as initializer-backed, never silently dropped by zip()."""
    declared = getattr(spec, "INITIALIZER_INPUTS", None)
    if declared is None:
        return set()
    if callable(declared):
        value = declared(framework, tc)
    elif isinstance(declared, dict):
        value = declared.get(framework, [])
    else:
        value = declared
    return set(value or [])


def assert_model_input_contract(framework, tc, spec, input_value_infos, inputs, initializer_inputs):
    model_inputs = _names_from_value_infos(input_value_infos)
    unexpected = sorted(set(initializer_inputs) - set(model_inputs))
    if unexpected:
        raise ValueError(
            f"op_spec INITIALIZER_INPUTS names not present in model inputs for {framework} tc{tc.get('id')}: {unexpected}")
    dynamic_inputs = [name for name in model_inputs if name not in initializer_inputs]
    if len(inputs) != len(dynamic_inputs):
        raise ValueError(
            f"op_spec make_inputs() returned {len(inputs)} arrays but model has {len(dynamic_inputs)} "
            f"dynamic inputs after INITIALIZER_INPUTS={sorted(initializer_inputs)} for {framework} "
            f"tc{tc.get('id')}. model_inputs={model_inputs}; dynamic_inputs={dynamic_inputs}")


_UNSET = object()


def assert_int8_genuine(spec, work_dir):
    """INT8 genuineness gate for the riscv_int8 path.

    A perfect INT8 cosine is meaningless if quantization silently BYPASSED the operator:
    if the op is missing from support_int8_ops_ (add-mslite ⑦) or its int8 OpCoder isn't
    registered (⑥), the tensors stay fp32, the FP32 opcoder is emitted, and every int8
    cosine prints a flat 1.000000 — green verdict, INT8 never exercised. SKILL.md used to
    flag "flat 1.000000" heuristically, but that also fires on legitimately discrete-output
    ops (hardmax/argmax/select) whose one-hot/index output IS identical under fp32 and int8.
    The real discriminator is not the cosine value but whether the int8 kernel was actually
    CALLED in the generated MCU code. Grep the codegen for the int8 kernel symbol.

    Returns (status: str, reason: str).
      * "genuine" when symbol is called in net*.c;
      * "missing" when symbol is absent (fp32 fallback or symbol misnamed);
      * "exempt" when INT8_KERNEL_SYMBOL == "" / None because this is a native integer/
        index operator with no fp32->int8 quantization path to prove.
    """
    symbol = getattr(spec, "INT8_KERNEL_SYMBOL", _UNSET)
    if symbol is _UNSET:
        symbol = f"{spec.OP_NAME}Int8"        # dominant nnacl convention (HardmaxInt8, ...)
    if not symbol:                            # explicitly disabled by the spec
        return "exempt", "native integer/index op: int8 genuineness check disabled (INT8_KERNEL_SYMBOL empty)"
    symbols = [symbol] if isinstance(symbol, str) else list(symbol)

    # The generated network code (where kernels are CALLED) lives at
    # <work_dir>/<fw>_riscv_micro/src/model*/net*.c. Search the call sites, NOT the copied
    # kernel-definition file (Collect() copies <op>_int8.c into the project, so the symbol's
    # *definition* may be present even on fp32 fallback — only the call in net*.c is proof).
    nets = sorted(work_dir.glob("**/net*.c"))
    if not nets:
        return "missing", (f"no generated net*.c under {work_dir} — codegen produced no network "
                           f"to inspect (build likely failed before emitting it)")
    blob = "\n".join(p.read_text(errors="ignore") for p in nets)
    for sym in symbols:
        if re.search(rf"\b{re.escape(sym)}\s*\(", blob):
            return "genuine", f"{sym}() called in codegen"
    return "missing", (
        f"INT8 kernel symbol {symbols} not called in generated code (searched {len(nets)} "
        f"net*.c) — quantization bypassed the op: the FP32 opcoder was emitted and this "
        f"path is an fp32 fallback, NOT real INT8 (the flat cos=1.0 trap). If the op's int8 "
        f"codegen genuinely uses a different name, declare INT8_KERNEL_SYMBOL in op_spec.py; "
        f"otherwise add the op to support_int8_ops_ and register its int8 OpCoder "
        f"(add-mslite ⑦ + ⑥).")


# --------------------------------------------------------------------------- #
# Per-case / per-framework runner
# --------------------------------------------------------------------------- #

def run_path(framework, path_key, case_dir, mslite_pkg, model_file,
             input_files, in_names, ref_outputs, spec):
    """Run one target path for one case. Returns (status, cos, msg).

    cos is None when the path could not run (build/convert error) -> FAIL.
    The cosine is computed HERE, in Python, from the benchmark's printed output tensors —
    the SAME way for x86 and riscv. No path lets the benchmark decide PASS/FAIL.
    """
    thr = PATH_META[path_key]["thr"]
    _, cfg_tmpl = DRIVER[(framework, path_key)]
    # Type-first classification: convert/ groups every path's *_micro build tree,
    # output/ groups every path's _run.sh / logs / device result. Each is split by path
    # underneath, so the three paths' conversion artifacts sit side by side (and so do
    # their outputs) rather than being scattered under per-path dirs.
    build_dir = case_dir / "convert" / path_key
    log_dir = case_dir / "output" / path_key

    # quant needs ONE calib dir PER input (different inputs -> different folders). Calib
    # data is a copy of the input bins, so it lives under input/ next to them.
    calib_dirs = []
    if path_key == "riscv_int8":
        for i, arr_path in enumerate(input_files):
            d = case_dir / "input" / f"calib_{i}"
            d.mkdir(exist_ok=True)
            (d / arr_path.name).write_bytes(arr_path.read_bytes())
            calib_dirs.append(d)
    # int8 cfg is materialized into convert/ (a conversion input); others use the template.
    cfg_file = make_cfg(path_key, cfg_tmpl, build_dir if path_key == "riscv_int8" else case_dir,
                        in_names, calib_dirs)

    stdout, stderr, rc = run_driver(
        framework, path_key, build_dir, log_dir, mslite_pkg, model_file, cfg_file, input_files,
        case_dir=case_dir, spec_path=getattr(spec, "__file__", None),
    )
    return judge_path_from_stdout(case_dir, path_key, stdout, stderr, rc, spec=spec, build_dir=build_dir)


def _err_msg(stdout, stderr, work_dir, rc=0):
    """Turn a failed path into an actionable one-line reason.

    Crashes (SIGABRT from heap corruption, SIGSEGV from a null/OOB deref) and timeouts are
    called out explicitly — these are the cases that used to read as a generic "no output"
    FAIL or, worse, hang the whole run. The data tensors' computation is what crashes, so the
    fix is in the operator's quant path (add-mslite side), not here.
    """
    blob = stdout + "\n" + stderr
    low = blob.lower()
    # 1) explicit crash / timeout from the return code — the most actionable signal
    if rc == -9 or "timed out after" in blob:
        return f"TIMEOUT — converter hung and was killed, likely heap corruption (see {work_dir}/stderr.log)"
    if rc and rc != 0:
        sig = -rc if rc < 0 else (rc - 128 if rc > 128 else None)  # bash reports 128+signal
        if sig in (6, 11):
            name = "SIGABRT — abort / heap corruption" if sig == 6 else "SIGSEGV — null/OOB deref"
            return f"converter crashed: {name} (see {work_dir}/stderr.log)"
    # 2) heap-corruption signatures in the logs even when the return code looks ordinary
    if "corrupted" in low or "sysmalloc" in low or "encounter an unknown error" in low:
        hit = next((l.strip() for l in blob.splitlines()
                    if any(s in l.lower() for s in ("corrupted", "sysmalloc", "unknown error"))), "")
        return f"converter crashed / heap corruption: {hit} (see {work_dir}/stderr.log)"
    # 3) generic build/convert error
    errs = [l for l in blob.splitlines() if "[ERR]" in l]
    head = errs[-1] if errs else (stderr.strip().splitlines()[-1] if stderr.strip() else "no benchmark output")
    return f"{head} (see {work_dir}/stderr.log)"


def run_framework(framework, active_paths, mslite_pkg, spec, project_dir, verbose):
    cases = spec.ONNX_TEST_CASES if framework == "onnx" else spec.TFLITE_TEST_CASES
    # Start from a clean slate: artifacts left by a previous run (possibly against an older
    # build) must never co-mingle with this run's — every number in the verdict has to come
    # from THIS run. Everything under output/<framework>/ is regenerable.
    fw_out = project_dir / "output" / framework
    if fw_out.exists():
        shutil.rmtree(fw_out)
    if framework == "tflite" and cases and getattr(spec, "TFLITE_TARGET_BUILTIN", None) is None:
        print("[warn] op_spec.py 未声明 TFLITE_TARGET_BUILTIN —— harness 无法核对每个模型确含目标 builtin。"
              "转换器会按形状把算子规范化成别的 builtin（用例静默测错对象），把存在性查证命中的 builtin 编号填进 spec。")
    rows = []
    for tc in cases:
        cid = tc.get("id", len(rows) + 1)
        case_dir = project_dir / "output" / framework / f"tc{cid}"
        case_dir.mkdir(parents=True, exist_ok=True)
        paths = {}
        try:
            # ----- Step 1: build model + inputs + reference (strict, gates the rest) -----
            # Case-level shared classification (computed once, reused by every target path):
            #   model/  the ONNX/TFLite model
            #   input/  input .bin files (+ int8 calib copies, added later in run_path)
            #   gt/     the onnxruntime/tf.lite reference outputs (ground truth), persisted
            model_dir = case_dir / "model"
            input_dir = case_dir / "input"
            gt_dir = case_dir / "gt"
            model_dir.mkdir(parents=True, exist_ok=True)
            input_dir.mkdir(parents=True, exist_ok=True)
            gt_dir.mkdir(parents=True, exist_ok=True)
            model_file = model_dir / f"model.{MODEL_EXT[framework]}"
            if framework == "onnx":
                spec.build_onnx_model(tc, str(model_file))
            else:
                spec.build_tflite_model(tc, str(model_file))
            _assert_target_builtin(spec, framework, model_file)
            raw_inputs = [np.asarray(a) for a in spec.make_inputs(tc, framework)]
            initializer_inputs = _initializer_inputs_for_case(spec, framework, tc)
            if framework == "onnx":
                import onnx
                model_proto = onnx.load(str(model_file))
                input_value_infos = list(model_proto.graph.input)
            else:
                import tensorflow as tf
                interp = tf.lite.Interpreter(model_path=str(model_file))
                input_value_infos = [type("TensorInfo", (), {"name": d["name"]}) for d in interp.get_input_details()]
            assert_model_input_contract(framework, tc, spec, input_value_infos, raw_inputs, initializer_inputs)
            inputs = raw_inputs
            in_names, out_names, ref_outputs = run_reference(framework, model_file, inputs)
            # Persist the reference outputs (ground truth) so gt/ is inspectable per case —
            # the harness still computes cosine from these in-memory; the files are for audit.
            for i, ref in enumerate(ref_outputs):
                np.save(gt_dir / (f"output_{i}.npy" if len(ref_outputs) > 1 else "output.npy"), ref)
            # ----- Step 4 (inputs persisted up-front; shared by all paths) -----
            input_files = []
            for i, arr in enumerate(inputs):
                p = input_dir / (f"input_{i}.bin" if len(inputs) > 1 else "input.bin")
                arr.tofile(p)
                input_files.append(p)
        except Exception as e:  # Step1 failure -> whole case FAIL, no path runs
            msg = f"Step1 build/reference failed: {e}"
            for pk in active_paths:
                paths[pk] = ("FAIL", None, msg)
            rows.append({"tc": tc, "paths": paths})
            _print_case(framework, cid, paths, active_paths)
            continue

        # ----- Step 2-5 per active target path (strict order, sequential) -----
        for pk in active_paths:
            status, cos, msg = run_path(
                framework, pk, case_dir, mslite_pkg, model_file,
                input_files, in_names, ref_outputs, spec,
            )
            paths[pk] = (status, cos, msg)
        rows.append({"tc": tc, "paths": paths})
        _print_case(framework, cid, paths, active_paths)

    excel = project_dir / f"{spec.OP_NAME}_{framework}_test_results.xlsx"
    write_excel(rows, excel, active_paths, framework, spec)
    print(f"[report] {excel}")
    return rows


def _print_case(framework, cid, paths, active_paths):
    cells = []
    for pk in active_paths:
        st, cos, _ = paths.get(pk, ("-", None, ""))
        cells.append(f"{pk}={'ERR' if cos is None else f'{cos:.4f}'}({st})")
    print(f"[{framework}] tc{cid}: " + "  ".join(cells))


# --------------------------------------------------------------------------- #
# Excel report (one file per framework)
# --------------------------------------------------------------------------- #

def write_excel(rows, excel_path, active_paths, framework, spec):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    blue = PatternFill("solid", fgColor="4472C4")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    white_bold = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{spec.OP_NAME} {framework}"[:31]

    cos_cols = [PATH_META[p]["col"] for p in active_paths]
    headers = ["用例编号", "描述"] + list(spec.PARAM_COLUMNS) + cos_cols + ["结果", "备注"]
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.alignment = blue, white_bold, center

    n_pass = 0
    for row in rows:
        tc = row["tc"]
        params = tc.get("params", {})
        line = [tc.get("id", ""), tc.get("desc", "")]
        line += [str(params.get(k, "")) for k in spec.PARAM_COLUMNS]
        cos_vals, ran_all, all_ok, notes = [], True, True, []
        for pk in active_paths:
            st, cos, msg = row["paths"].get(pk, ("-", None, ""))
            cos_vals.append(cos)
            if cos is None:
                ran_all = False
                all_ok = False
            elif st != "PASS":
                all_ok = False
            if msg:
                notes.append(f"{pk}:{msg}")
        passed = ran_all and all_ok
        n_pass += int(passed)
        line += [("" if c is None else round(c, 6)) for c in cos_vals]
        line += ["PASS" if passed else "FAIL", "; ".join(notes)]
        ws.append(line)

        r = ws.max_row
        fill = green if passed else red
        for c in ws[r]:
            c.fill = fill
        # number format for cosine columns
        for j in range(len(cos_vals)):
            col = 2 + len(spec.PARAM_COLUMNS) + 1 + j
            ws.cell(row=r, column=col).number_format = "0.000000"

    total = len(rows)
    thr_desc = ", ".join(f"{PATH_META[p]['col']}>={PATH_META[p]['thr']}" for p in active_paths)
    ws.append([])
    ws.append([f"总计 {total}", f"通过 {n_pass}", f"失败 {total - n_pass}"]
              + [""] * (len(headers) - 3))
    ws.append([f"判定标准: {thr_desc} (全部满足判 PASS)"] + [""] * (len(headers) - 1))

    wb.save(excel_path)


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #

# --------------------------------------------------------------------------- #
# Anti-shrink gate: refuse to run when the case set regressed vs the last round
# --------------------------------------------------------------------------- #

_SUMMARY_LINE_RE = re.compile(r"^(\w+)\s+tc(\S+)\s+(\S+)\s+(PASS|FAIL)\b")


def check_case_regression(prev_summary: Path, spec, frameworks):
    """防"删用例换绿"闸门：本轮用例集相比上一轮 verify_summary.txt 缩水 → 拒绝开跑。

    历史事故：广播用例实跑 FAIL 后被从 op_spec 删除，重跑得到 "0 FAIL" 当作完成上报。
    VERDICT 的分母是能力覆盖，不是 op_spec 现存用例——删用例不会让缺陷消失，只会让结论失效。
    确属有意缩减（须经用户裁决、汇报中列为覆盖缺口）时，OP_VERIFY_ACK_REDUCED=1 显式放行，
    豁免行会写进本轮 VERDICT 留痕。返回值：要附进 verdict 的留痕行（无缩水 / 未 ack 时为空）。
    """
    if not prev_summary.is_file():
        return []
    prev_ids, prev_fail_ids, prev_op = {}, {}, None
    for line in prev_summary.read_text().splitlines():
        if prev_op is None and line.startswith("op="):
            prev_op = line.split()[0][3:]
        m = _SUMMARY_LINE_RE.match(line)
        if m:
            fw, cid, _, st = m.groups()
            prev_ids.setdefault(fw, set()).add(cid)
            if st == "FAIL":
                prev_fail_ids.setdefault(fw, set()).add(cid)
    if prev_op != spec.OP_NAME:   # 不同算子的残留 summary，不可比
        return []
    cur_ids = {"onnx": {str(tc.get("id", "")) for tc in spec.ONNX_TEST_CASES},
               "tflite": {str(tc.get("id", "")) for tc in spec.TFLITE_TEST_CASES}}
    warnings = []
    for fw in frameworks:
        if fw not in prev_ids:
            continue
        missing = sorted(prev_ids[fw] - cur_ids[fw], key=str)
        if missing:
            gone_fail = sorted(prev_fail_ids.get(fw, set()) - cur_ids[fw], key=str)
            warnings.append(
                f"CASES_REDUCED [{fw}] {len(prev_ids[fw])}→{len(cur_ids[fw])} "
                f"上轮存在、本轮缺席: tc{',tc'.join(missing)}"
                + (f"（其中上轮 FAIL: tc{',tc'.join(gone_fail)}）" if gone_fail else ""))
    if not warnings:
        return []
    if os.environ.get("OP_VERIFY_ACK_REDUCED") == "1":
        return ["ACK_REDUCED（已显式确认缩减——汇报中必须把缺席用例列为覆盖缺口）:"] \
               + ["  " + w for w in warnings]
    sys.exit(
        "[ERROR] 用例集相比上一轮缩水，拒绝开跑（防\"删 FAIL 用例换绿\"闸门）。\n"
        + "".join(f"        {w}\n" for w in warnings)
        + "        FAIL 用例只有两个合法出路：修实现代码，或用日志级证据证明用例本身设计错误。\n"
        "        删除/缺席用例得到的 0 FAIL 不构成完成。确属用户裁决的有意缩减（如该形态经实证\n"
        "        无法以目标 builtin 产出，列为覆盖缺口）时，OP_VERIFY_ACK_REDUCED=1 显式放行，\n"
        "        豁免将记入本轮 VERDICT。")


# --------------------------------------------------------------------------- #
# Capability-coverage gate: the add-mslite step3 checklist, frozen as JSON, must stay covered
# --------------------------------------------------------------------------- #

def load_capability_checklist(project_dir: Path):
    """add-mslite step3 产物三「能力验收清单」frozen as JSON at
    <project>/scripts/capability_checklist.json. Optional, but ENFORCED when present.

    Defends the real failure mode this gate was built for: the step3 checklist (forms/axes/
    attributes the op committed to support) being silently REWRITTEN at backfill time to
    match whatever test cases already existed — coverage looks complete while a committed
    form was never tested. Freezing the checklist as a machine-read artifact means the
    VERDICT reports coverage against the COMMITMENT, not against the prose the model emits.
    Returns the parsed dict, or None when absent (then coverage is merely warned)."""
    path = project_dir / "scripts" / "capability_checklist.json"
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text())
    except Exception as e:
        sys.exit(f"[ERROR] capability_checklist.json 解析失败: {e}")
    caps = data.get("capabilities")
    if not isinstance(caps, list) or not caps:
        sys.exit("[ERROR] capability_checklist.json 必须含非空 'capabilities' 数组"
                 "（每条 {id, desc, covered_by:[case_id,...], match?:{param:value}}）。")
    return data


def _spec_case_params(spec, frameworks, case_id):
    """Return the params dict of the test case with this id (first hit across frameworks)."""
    for fw in frameworks:
        cases = spec.ONNX_TEST_CASES if fw == "onnx" else spec.TFLITE_TEST_CASES
        for tc in cases:
            if tc.get("id") == case_id:
                return tc.get("params", {})
    return None


def validate_checklist_refs(checklist, spec, frameworks):
    """PRE-run structural gate (runs before burning a 10-min round):
      * every capability has a NON-EMPTY covered_by (a committed form with no case = an
        uncovered commitment — exactly the silent gap this defends against);
      * every covered_by id EXISTS in an in-scope framework's case set (no dangling refs);
      * if a capability declares an optional `match` params predicate, at least one of its
        covered_by cases must actually have those params — so covered_by can't point at an
        unrelated case to fake coverage (lightweight defence; full rigor is opt-in)."""
    all_ids = set()
    for fw in frameworks:
        cases = spec.ONNX_TEST_CASES if fw == "onnx" else spec.TFLITE_TEST_CASES
        all_ids |= {tc.get("id") for tc in cases}
    problems = []
    for cap in checklist["capabilities"]:
        cid = cap.get("id", "?")
        cov = cap.get("covered_by") or []
        if not cov:
            problems.append(f"能力 {cid}「{cap.get('desc','')}」无 covered_by 用例 —— "
                            f"承诺支持却无验证用例（不得删能力行，须补用例）")
            continue
        missing = [i for i in cov if i not in all_ids]
        if missing:
            problems.append(f"能力 {cid} 的 covered_by={cov} 引用不存在的用例 id {missing}")
        match = cap.get("match")
        if match and not any(
                (p := _spec_case_params(spec, frameworks, i)) is not None
                and all(p.get(k) == v for k, v in match.items()) for i in cov):
            problems.append(f"能力 {cid} 声明 match={match}，但其 covered_by 用例无一 params 匹配 "
                            f"—— covered_by 可能指错了用例")
    if problems:
        sys.exit("[ERROR] capability_checklist.json 对账失败（开跑前拦截，防能力清单被悄悄改写）:\n"
                 + "".join(f"        - {p}\n" for p in problems)
                 + "        合法处置：补齐/修正用例的 covered_by，或为某能力补实测用例。\n"
                 "        对账方向单向——以 add-mslite step3 能力清单为准改 spec，不得反向删/改能力行来匹配存量 spec。")


def report_capability_coverage(checklist, passed_case_ids):
    """POST-run: fold the frozen checklist into the verdict. Each capability is covered iff
    ≥1 of its covered_by cases PASSED this round. Returns (lines, n_uncovered) — uncovered
    capabilities make the run non-green (a committed form whose cases all failed/skipped is
    a real gap, regardless of the raw case tally)."""
    lines, uncovered = [], 0
    for cap in checklist["capabilities"]:
        cov = cap.get("covered_by") or []
        hit = [i for i in cov if i in passed_case_ids]
        if hit:
            lines.append(f"  [COVERED]   {cap.get('id','?')}: {cap.get('desc','')} "
                         f"← tc{',tc'.join(str(i) for i in hit)}")
        else:
            uncovered += 1
            lines.append(f"  [UNCOVERED] {cap.get('id','?')}: {cap.get('desc','')} "
                         f"(covered_by={cov} 无一通过)")
    head = f"CAPABILITY COVERAGE: {len(checklist['capabilities']) - uncovered}/" \
           f"{len(checklist['capabilities'])} 能力被通过用例覆盖" \
           + (f"，{uncovered} 条未覆盖" if uncovered else "")
    return [head, *lines], uncovered


def main():
    ap = argparse.ArgumentParser(description="hs-verify-op fixed harness")
    ap.add_argument("--spec", default="scripts/op_spec.py",
                    help="算子 spec 文件路径 (默认 scripts/op_spec.py)")
    ap.add_argument("--framework", choices=["onnx", "tflite", "all"], default="all")
    ap.add_argument("--target", choices=["x86", "riscv", "all"], default="all")
    ap.add_argument("--threshold-fp32", type=float, default=DEFAULT_THRESHOLD_FP32,
                    help="调试用：放宽 fp32 阈值（绝不用于结论性签收）")
    ap.add_argument("--threshold-int8", type=float, default=DEFAULT_THRESHOLD_INT8,
                    help="调试用：放宽 INT8 阈值（绝不用于结论性签收）")
    ap.add_argument("--judge-case", help="手动重跑后刷新并判定单个 case 目录")
    ap.add_argument("--judge-path", choices=["x86_fp32", "riscv_fp32", "riscv_int8"],
                    help="--judge-case 对应的目标路径")
    ap.add_argument("--stdout-log", help="--judge-case 使用的最新 stdout.log")
    ap.add_argument("--stderr-log", help="--judge-case 使用的最新 stderr.log")
    ap.add_argument("--driver-rc", type=int, default=0,
                    help="--judge-case 对应 _run.sh driver 退出码")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    PATH_META["x86_fp32"]["thr"] = args.threshold_fp32
    PATH_META["riscv_fp32"]["thr"] = args.threshold_fp32
    PATH_META["riscv_int8"]["thr"] = args.threshold_int8

    if args.judge_case:
        if not args.judge_path:
            sys.exit("[ERROR] --judge-case 需要同时指定 --judge-path")
        case_dir = Path(args.judge_case).resolve()
        log_dir = case_dir / "output" / args.judge_path
        stdout_log = Path(args.stdout_log).resolve() if args.stdout_log else log_dir / "stdout.log"
        stderr_log = Path(args.stderr_log).resolve() if args.stderr_log else log_dir / "stderr.log"
        if not stdout_log.is_file():
            sys.exit(f"[ERROR] stdout log 不存在: {stdout_log}")
        stdout = stdout_log.read_text()
        stderr = stderr_log.read_text() if stderr_log.is_file() else ""
        spec = None
        spec_path_arg = Path(args.spec).resolve() if args.spec and args.spec != "scripts/op_spec.py" else None
        if spec_path_arg and spec_path_arg.is_file():
            spec = load_spec_module(spec_path_arg)
        build_dir = case_dir / "convert" / args.judge_path
        status, cos, msg = judge_path_from_stdout(
            case_dir, args.judge_path, stdout, stderr, args.driver_rc, spec=spec, build_dir=build_dir,
        )
        report = (log_dir / "judge.txt").read_text().strip()
        print(f"[judge] {report}")
        sys.exit(0 if status == "PASS" else 1)

    # Anchor project_dir to the SPEC location, not the CWD. A re-run launched without
    # `cd` once scattered output/ + reports into the wrong directory and silently burned
    # a full 10+ min round (a real past mistake) — CWD is no longer a semantic input.
    # Convention: spec lives at <project>/scripts/op_spec.py → project_dir = <project>.
    spec_path = (Path(args.spec) if os.path.isabs(args.spec) else Path.cwd() / args.spec).resolve()
    if not spec_path.is_file():
        sys.exit(f"[ERROR] spec 文件不存在: {spec_path}\n"
                 "        约定: <算子项目目录>/scripts/op_spec.py（相对 --spec 按当前目录解析，"
                 "推荐直接传绝对路径）")
    project_dir = spec_path.parent.parent if spec_path.parent.name == "scripts" else spec_path.parent

    # Guard: NEVER anchor inside the MindSpore Lite source/build tree. The harness writes
    # output/ and <op>_<fw>_test_results.xlsx + verify_summary.txt into project_dir;
    # anchoring in the submodule pollutes it (a real past mistake). Keep op_spec.py in the
    # operator project dir (convention: $MSLITE_OP_OUTPUT/<op>/scripts/op_spec.py;
    # default beside the mindspore-lite repo, i.e. src/mslite-op-output/<op> inside HiSpark.AI).
    if (project_dir / "schema" / "ops.fbs").is_file() or \
       (project_dir / "mindspore-lite" / "schema" / "ops.fbs").is_file():
        sys.exit(
            "[ERROR] 由 --spec 推导出的项目目录落在 MindSpore Lite 源码/构建树内。\n"
            f"        project_dir: {project_dir}\n"
            "        请把 op_spec.py 放到算子项目目录(约定 $MSLITE_OP_OUTPUT/<op>/scripts/，缺省与 mindspore-lite 仓平级，HiSpark.AI 仓内即 src/mslite-op-output/<op>/scripts/)；\n"
            "        报告(verify_summary.txt / *_test_results.xlsx)与 output/ 都写在 project_dir，\n"
            "        否则会污染 mindspore-lite submodule。"
        )
    if project_dir != Path.cwd().resolve():
        print(f"[env] project_dir 由 --spec 推导（与当前工作目录无关）: {project_dir}")

    frameworks = FRAMEWORKS[args.framework]
    if "onnx" in frameworks:
        ensure("onnxruntime")
    if "tflite" in frameworks:
        ensure("tensorflow")
    ensure("openpyxl")

    spec = load_spec(spec_path)
    # 开跑前对账上一轮 summary：用例集缩水 = 结论分母被偷换，先拦截再谈跑。
    ack_lines = check_case_regression(project_dir / "verify_summary.txt", spec, frameworks)
    # 开跑前对账能力清单：每条承诺的能力须有存在的 covered_by 用例（防清单被悄悄改写/缩减）。
    checklist = load_capability_checklist(project_dir)
    if checklist:
        validate_checklist_refs(checklist, spec, frameworks)
    else:
        print("[warn] 未发现 scripts/capability_checklist.json —— 能力覆盖无法机械校验。"
              "add-mslite step3 应落盘该清单；缺失时 VERDICT 不含能力覆盖留痕。")
    mslite_pkg = resolve_mslite_pkg(project_dir)
    active_paths = TARGET_PATHS[args.target]
    # Traceability: which converter_lite produced this verdict (build time = archive mtime).
    conv_built = datetime.datetime.fromtimestamp(
        Path(mslite_pkg, "tools/converter/converter/converter_lite").stat().st_mtime
    ).strftime("%Y-%m-%d %H:%M:%S")
    print(f"[env] MSLITE_PKG = {mslite_pkg}")
    print(f"[env] converter_lite built at {conv_built}")
    print(f"[env] project    = {project_dir}")
    print(f"[run] op={spec.OP_NAME} frameworks={frameworks} paths={active_paths}")

    # Run every framework, then emit a single machine-checkable verdict. The verdict
    # (and the per-framework Excel) are the ONLY trustworthy result surface: report it
    # verbatim, never from memory. Exit code mirrors it (nonzero == at least one FAIL).
    summary_lines = []
    total = passed = 0
    passed_case_ids = set()   # ids whose case passed ALL active paths in ≥1 framework
    for fw in frameworks:
        rows = run_framework(fw, active_paths, mslite_pkg, spec, project_dir, args.verbose)
        for row in rows:
            cid = row["tc"].get("id", "")
            row_ok = True
            for pk in active_paths:
                st, cos, msg = row["paths"].get(pk, ("FAIL", None, "not run"))
                ok = (cos is not None and st == "PASS")
                row_ok = row_ok and ok
                total += 1
                passed += int(ok)
                cos_s = "ERR" if cos is None else f"{cos:.6f}"
                summary_lines.append(
                    f"{fw:<6} tc{cid:<3} {pk:<11} {st:<4} cos={cos_s}"
                    + (f"  {msg}" if msg else ""))
            if row_ok:
                passed_case_ids.add(cid)

    # Fold the frozen capability checklist into the verdict: a committed form whose covering
    # cases all failed is an uncovered commitment, which makes the run non-green even if the
    # raw tally looks fine. The lines echo the checklist verbatim — coverage is reported
    # against the COMMITMENT, not the model's prose.
    cap_lines, cap_uncovered = ([], 0)
    if checklist:
        cap_lines, cap_uncovered = report_capability_coverage(checklist, passed_case_ids)

    failed = total - passed
    exit_code = 1 if (failed or cap_uncovered) else 0
    cap_tag = ""
    if checklist:
        cap_tag = (f"  capabilities={len(checklist['capabilities']) - cap_uncovered}/"
                   f"{len(checklist['capabilities'])}")
    verdict = (f"VERDICT: op={spec.OP_NAME}  {passed}/{total} variant-cases PASS, "
               f"{failed} FAIL{cap_tag}  "
               f"thresholds(fp32={args.threshold_fp32}, int8={args.threshold_int8})")
    header = (f"hs-verify-op summary\nop={spec.OP_NAME}  frameworks={frameworks}  "
              f"paths={active_paths}\nMSLITE_PKG={mslite_pkg}\n"
              f"converter_lite built at {conv_built}\n" + "-" * 60)
    cap_block = (["-" * 60, *cap_lines] if cap_lines else [])
    # HARNESS_EXIT 紧跟 VERDICT 写入日志与 summary：nohup 后台模式下进程退出码不可观测，
    # 这一行是调用方唯一可靠的退出码来源（自行 grep FAIL 计数会把 "0 FAIL" 也算成失败）。
    report = "\n".join([header, *summary_lines, *cap_block, "-" * 60, *ack_lines, verdict,
                        f"HARNESS_EXIT={exit_code}"])
    (project_dir / "verify_summary.txt").write_text(report + "\n")
    print("\n" + report)
    print(f"[summary] {project_dir / 'verify_summary.txt'}")

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
